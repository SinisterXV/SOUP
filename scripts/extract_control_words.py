from openpyxl import load_workbook
import sys

# Class to handle a sheet of a xlsx file
class Sheet:
    # Opens the sheet
    def __init__(self, file_path):
        self.cw_file = load_workbook(filename=file_path)
        self.cw_sheet = self.cw_file.active

    # Access a single cell of the file
    def access_value(self, row, column):
        return self.cw_sheet.cell(row = row, column = column).value

def main():     
    # Opens the file 
    try:
        S = Sheet(sys.argv[1])
    except:
        print("Provide the correct file to open as second argument")
        exit()

    # Rows and column are addressed starting from one: we skip the first row, in which there 
    # are the headings of the columns
    row_instruction = 2

    # For each row in the file, we print the the control word together with the mnemonic information
    while S.access_value(row_instruction, 1) != None:
        string_to_print=""
        # Read mnemonic
        instruction_mnemonic = S.access_value(row_instruction, 1)
        # Read opcode
        instruction_opcode = S.access_value(row_instruction, 2)
        # Write mnemonic, opcode in dec and opcode in hex
        string_to_print = ''.join([string_to_print, f"{instruction_mnemonic}\t"])
        string_to_print = ''.join([string_to_print, f"{instruction_opcode}\t"])
        string_to_print = ''.join([string_to_print, f"{hex(int(instruction_opcode))}\t"])

        # Read control word
        control_word = ""
        column_instruction = 3
        # For each column        
        while(S.access_value(row_instruction, column_instruction) != None):
            # Read the content of the cell
            to_add = str(S.access_value(row_instruction, column_instruction))
            # If one value is X, it is replaces with 0
            to_add = to_add.replace('X','0')
            control_word = ''.join([control_word, to_add])
            # Next column
            column_instruction += 1

        # Write the control word
        string_to_print = ''.join([string_to_print, f"{control_word}\t"])
        # Print the line
        print(string_to_print)
        # Next row
        row_instruction += 1
        
if __name__ == "__main__":
    main()